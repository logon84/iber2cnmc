#!/usr/bin/python3 
import requests
import openpyxl
import sys, getopt
import os

CUPS = '6666666F666666660F'
daynames = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
monthnames = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

def get_xy_cell (row, col):
	global sheet
	val = str(sheet.cell(row = row, column = col).value)
	if val == 'None':
		return ''
	else:
		return val

		
def get_int_hour (s): #remove starting zeros & minutes from hours
	if s[:1] == '0':
		return s[1:2] #if hour starts with zero, remove it at return and also minutes
	else:
		return s[:2] #just remove minutes

def set_std_dayformat (s):
	if len(s) == 1:
		s = '0'+s #if day is  just one char, put a zero before to ensure dd format

	head = get_xy_cell (1, 1) #get xlsx title
	year = head.split(' ')
	year = year[len(year)-1] #get year from title
	if " del " in head:
		print("El formato de entrada es SEMANAL y no está soportado.")
		sys.exit(0)
	if any(day in head for day in daynames):
		print("El formato de entrada es DIARIO y no está soportado.")
		sys.exit(0)
	if any(month in head for month in monthnames):
		for month in monthnames:
			if month in head:
				if (1+monthnames.index(month)) <10:
					return s + '/0' + str(1+(monthnames.index(month))) + '/' + year #ensure moth in mm format
				else:
					return s + '/' + str(1+(monthnames.index(month))) + '/' + year

def xlsx2csv(inname, outname):
	global sheet
	csv_head = 'CUPS;DIA;Hora;Mi consumo;Metodo Obtencion\n'
	out_file = csv_head
	book = openpyxl.load_workbook(inname) #open xlsx input file
	sheet = book.active
	rows = sheet.rows
	for row in range (4, sheet.max_row+1):
		date = set_std_dayformat(get_xy_cell (row, 1))
		hour = get_int_hour(get_xy_cell (row, 2))
		consumption = str(float(get_xy_cell (row, 3))/1000) #convert watts*h to kwatts*h
		consumption = consumption.replace('.',',')#replace float separator
		obtained = "R"
		out_file = out_file + CUPS + ';' + date + ';' + hour + ';' + consumption + ';' + obtained + "\n"
	file = open(outname,'w')
	file.write (out_file)
	file.close()

def main(argv):
   inputfile = ''
   outputfile = ''
   try:
      opts, args = getopt.getopt(argv,"hi:o:",["ifile=","ofile="])
   except getopt.GetoptError:
      print('iber2cnmc.py -i <inputfile.xlsx> -o <outputfile.csv>')
      sys.exit(2)
   for opt, arg in opts:
      if opt == '-h':
         print('iber2cnmc.py -i <inputfile.xlsx> -o <outputfile.csv>')
         sys.exit()
      elif opt in ("-i", "--ifile"):
         inputfile = arg
      elif opt in ("-o", "--ofile"):
         outputfile = arg
   xlsx2csv(inputfile, outputfile)

if __name__ == "__main__":
   main(sys.argv[1:])